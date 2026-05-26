import pandas as pd
import io
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_bytes(data: list) -> tuple:
    """
    Generates a beautified Excel file from parsed VSS-110 reports data.
    
    Args:
        data (list): A list of parsed report records.
        
    Returns:
        tuple: (bytes, str) -> The Excel file bytes and the formatted filename.
    """
    # Create DataFrame with specific column order as requested
    columns = [
        "REPORT_ID",
        "REPORTING_FOR",
        "REPORT_DATE",
        "SETTLEMENT_CURRENCY",
        "REIMBURSEMENT_FEES_CREDIT",
        "REIMBURSEMENT_FEES_DEBIT",
        "REIMBURSEMENT_FEES_TOTAL",
        "VISA_CHARGES_CREDIT",
        "VISA_CHARGES_DEBIT",
        "VISA_CHARGES_TOTAL"
    ]
    
    df = pd.DataFrame(data, columns=columns)
    
    # Generate timestamp for output name
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"VSS110_Output_{timestamp}.xlsx"
    
    # Write to memory buffer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="VSS-110 Reports")
        
        # Access openpyxl sheet
        workbook = writer.book
        worksheet = writer.sheets["VSS-110 Reports"]
        
        # UI Styling Elements
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Classic Navy Header
        font_data = Font(name="Segoe UI", size=10)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_border_side = Side(border_style="thin", color="D9D9D9")
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Style Header Row (Row 1)
        worksheet.row_dimensions[1].height = 26
        for col_idx in range(1, len(columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = cell_border
            
        # Style Data Rows
        num_rows = len(df)
        for r_idx in range(2, num_rows + 2):
            worksheet.row_dimensions[r_idx].height = 20
            for c_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.font = font_data
                cell.border = cell_border
                
                col_name = columns[c_idx - 1]
                
                # Alignments and Number Formats based on data type
                if col_name in ["REPORT_ID", "REPORT_DATE", "SETTLEMENT_CURRENCY"]:
                    cell.alignment = align_center
                elif col_name in ["REPORTING_FOR"]:
                    cell.alignment = align_left
                else:
                    # Currency numeric/financial text values
                    cell.alignment = align_right
                    
        # Auto-adjust column widths with generous padding
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Use cell length after string conversion
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Add safety padding
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Seek to start
    output.seek(0)
    return output.read(), filename
